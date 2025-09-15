import logging
from datetime import datetime
import numpy as np
import textwrap
import os
import tempfile
import xlwings as xw
import pandas as pd
from pathlib import Path
import re
from openpyxl import load_workbook


def setup_logger():
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)

    if not logger.handlers:
        # Create logs/ directory if it doesn't exist#
        log_dir = resolve_path_relative_to_script("./logs")

        os.makedirs(log_dir, exist_ok=True)

        # Create a unique log file with timestamp down to the second
        log_file = os.path.join(log_dir, f"log_{datetime.now():%Y%m%d_%H%M%S}.log")

        # File handler
        fh = logging.FileHandler(log_file, mode='w', encoding='utf-8')
        fh.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)
        logger.addHandler(fh)

        # Optional: Console handler
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        ch.setFormatter(formatter)
        logger.addHandler(ch)

    return logger


def set_dropdown_values(workbook_name, sheet_name, dropdown_name, items):
    """
    Modifies a Form Control dropdown (not Data Validation) by its name in a specific open workbook.

    Parameters:
        workbook_name (str): Name of the open Excel workbook (e.g., "Budget2025.xlsx").
        sheet_name (str): Name of the sheet containing the dropdown.
        dropdown_name (str): Name of the Form Control dropdown (e.g., "Drop Down 6").
        items (list of str): List of options to populate the dropdown with.

    Returns:
        True if successful, False otherwise.
    """

    try:
        for app in xw.apps:
            for wb in app.books:
                if wb.name == workbook_name:
                    try:
                        sheet = wb.sheets[sheet_name]
                        dropdown = sheet.api.Shapes(dropdown_name).ControlFormat
                        dropdown.RemoveAllItems()
                        for item in items:
                            dropdown.AddItem(item)
                        return True
                    except Exception as e:
                        print(f"Error accessing dropdown in {workbook_name}: {e}")
                        continue
    except Exception as e:
        print(f"Error while setting dropdown values: {e}")

    print(f"Workbook '{workbook_name}' or dropdown '{dropdown_name}' not found.")
    return False


def resolve_path_relative_to_script(path):
    """
    Resolves a given path relative to the script's directory.
    If the path is already absolute, it is returned unchanged.

    Parameters:
    - path: str, the path to resolve

    Returns:
    - str, the absolute path resolved relative to this script
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    return path if os.path.isabs(path) else os.path.abspath(os.path.join(script_dir, path))


def write_df(workbook_name, sheet_name, upper_left_address, dataframe, include_headers=True):
    """
     Writes a Pandas DataFrame to an already open Excel workbook, starting at a given cell address or named range.

     Parameters:
     workbook_name (str): Name of the open Excel workbook (e.g. 'GeometrieConverter.xlsx').
     sheet_name (str): Name of the sheet within the workbook.
     upper_left_address (str): Excel address (e.g., 'A1') or named range where the DataFrame should be placed.
     dataframe (pd.DataFrame): The Pandas DataFrame to write to the Excel sheet.
     include_headers (bool): Whether to include column names as headers in Excel (default is True).
     """

    try:
        # Connect to the already open workbook
        wb = xw.books[workbook_name]
        sheet = wb.sheets[sheet_name]

        # Check if upper_left_address is a named range
        try:
            upper_left_range = wb.names[upper_left_address].refers_to_range
        except KeyError:
            # Not a named range, treat it as a cell address
            upper_left_range = sheet.range(upper_left_address)

        # Prepare data
        if include_headers:
            data = [dataframe.columns.tolist()] + dataframe.values.tolist()
        else:
            data = dataframe.values.tolist()

        # Write data
        upper_left_range.value = data

    except Exception as e:
        print(f"Error writing DataFrame to Excel: {e}.")


def write_value(workbook_name, sheet_name, cell_or_named_range, value):
    """
    Writes a single value to a specific cell or named range in an already open Excel workbook.

    Parameters:
    workbook_name (str): Name of the open Excel workbook (e.g. 'GeometrieConverter.xlsx').
    sheet_name (str): Name of the sheet within the workbook.
    cell_or_named_range (str): Excel address (e.g., 'B2') or a named range where the value should be written.
    value (any): The value to write into the cell.
    """

    try:
        # Connect to the already open workbook
        wb = xw.books[workbook_name]
        sheet = wb.sheets[sheet_name]

        # Try to resolve the named range or cell address
        try:
            target_range = wb.names[cell_or_named_range].refers_to_range
        except KeyError:
            target_range = sheet.range(cell_or_named_range)

        # Write the single value
        target_range.value = value

    except Exception as e:
        print(f"Error writing value to Excel: {e}.")


def write_df_to_table(workbook_name, sheet_name, table_name, dataframe):
    """
    Replace the contents of an existing Excel table with a pandas DataFrame using xlwings.

    If the DataFrame is empty, the table is cleared but not resized or filled with rows.

    Parameters:
    - workbook_name: str, name of the open Excel workbook (no path needed if open).
    - sheet_name: str, name of the sheet containing the table.
    - table_name: str, name of the Excel table (ListObject) to manipulate.
    - dataframe: pandas DataFrame to replace the table's data (same number of columns).
    """
    # Connect to the open workbook
    wb = xw.books[workbook_name]
    ws = wb.sheets[sheet_name]

    # Find the table (ListObject)
    try:
        table = ws.api.ListObjects(table_name)
    except Exception as e:
        raise ValueError(f"Table '{table_name}' not found in sheet '{sheet_name}'.") from e

    # Get the header range and data body range
    header_range = table.HeaderRowRange

    try:
        data_body_range = table.DataBodyRange
        if data_body_range:
            data_body_range.ClearContents()
    except Exception as e:
        print("No DataBodyRange to clear:", e)


    # If the DataFrame is empty, return early after clearing
    if dataframe.empty:
        return

    # Clean DataFrame to avoid writing the index
    df_clean = dataframe.reset_index(drop=True)

    # Write the new DataFrame below the headers
    start_cell = ws.range((header_range.Row + 1, header_range.Column))
    start_cell.options(index=False, header=False).value = df_clean

    # Resize the table to match the new data
    last_row = header_range.Row + df_clean.shape[0]
    last_col = header_range.Column + df_clean.shape[1] - 1
    new_range = ws.range(
        (header_range.Row, header_range.Column),
        (last_row, last_col)
    )
    table.Resize(new_range.api)


def show_message_box(workbook_name, message, buttons="vbOK", icon="vbInformation",
                     default="vbDefaultButton1", title="Message"):
    """
    Shows a message box in Excel by injecting VBA code dynamically and calling it.

    Args:
        workbook_name (str): The name of the open Excel workbook (e.g., "Book1.xlsm").
        message (str): The message to show in the message box.
        buttons (str): VBA button constant (e.g., "vbYesNo", "vbOKCancel", etc.).
        icon (str): VBA icon constant (e.g., "vbExclamation", "vbInformation", "vbCritical").
        default (str): VBA default button constant (e.g., "vbDefaultButton2").
        title (str): Title of the message box.

    Returns:
        str: The caption of the clicked button (e.g., "Yes", "No", "Cancel"), or None on failure.
    """

    VBA_BUTTONS = {
        "vbOK": 0,
        "vbOKCancel": 1,
        "vbAbortRetryIgnore": 2,
        "vbYesNoCancel": 3,
        "vbYesNo": 4,
        "vbRetryCancel": 5
    }

    VBA_ICONS = {
        "vbCritical": 16,
        "vbQuestion": 32,
        "vbExclamation": 48,
        "vbInformation": 64
    }

    VBA_DEFAULTS = {
        "vbDefaultButton1": 0,
        "vbDefaultButton2": 256,
        "vbDefaultButton3": 512,
        "vbDefaultButton4": 768
    }

    response_map = {
        1: "OK",
        2: "Cancel",
        3: "Abort",
        4: "Retry",
        5: "Ignore",
        6: "Yes",
        7: "No"
    }

    if not isinstance(message, str):
        raise TypeError(f"Expected message to be a string, but got {type(message).__name__}")

    app = xw.apps.active
    wb = app.books[workbook_name]

    msgbox_flags = VBA_BUTTONS.get(buttons, 0) + VBA_ICONS.get(icon, 0) + VBA_DEFAULTS.get(default, 0)

    # --- Prepare message for VBA ---
    message_escaped = (
        message.replace('"', '""')  # Escape quotes
        .replace('\t', '" & vbTab & "')  # Replace tabs
        .replace('\n', '" & vbNewLine & "')  # Replace newlines
    )

    # Split into safe chunks to avoid line overflow in VBA
    max_chunk_length = 900
    chunks = textwrap.wrap(str(message_escaped), width=max_chunk_length)
    message_vba = ' & _\n    '.join(f'"{chunk}"' for chunk in chunks)

    # --- VBA code injection ---
    vba_code = f"""
    Function ShowMessageBox() As Integer
        ShowMessageBox = MsgBox({message_vba}, {msgbox_flags}, "{title}")
    End Function
    """

    module_name = "MsgBoxTemp"
    vbproj = wb.api.VBProject

    # Add or get the module
    try:
        vb_module = vbproj.VBComponents(module_name)
    except Exception:
        vb_module = vbproj.VBComponents.Add(1)  # 1 = standard module
        vb_module.Name = module_name

    code_module = vb_module.CodeModule

    # Clear old code and insert new
    count_lines = code_module.CountOfLines
    if count_lines > 0:
        code_module.DeleteLines(1, count_lines)
    code_module.AddFromString(vba_code)

    # Run the function
    result = wb.macro("ShowMessageBox")()

    return response_map.get(result, f"Unknown ({result})")


def read_excel_table(workbook_name, sheet_name, table_name, dtype=None, dropnan=False, strip=True):
    """
    Read an Excel Table into a Pandas DataFrame, using the Table's header as column names.

    Parameters:
        workbook_name (str): The name of the workbook
        sheet_name (str): The name of the sheet containing the table.
        table_name (str): The name of the Excel Table (not the range name).
        dtype (type or dict): Optional dtype to cast columns to.
        dropnan (bool): Whether to drop rows that are completely NaN.
        strip (bool): Whether to strip whitespace from string cells.

    Returns:
        pd.DataFrame: DataFrame containing the table data with correct headers.
    """

    wb = xw.Book(workbook_name)
    sheet = wb.sheets[sheet_name]
    table = sheet.tables[table_name]

    data_range = table.data_body_range
    headers = [h.strip() if isinstance(h, str) else str(h) for h in table.header_row_range.value]

    if data_range is None:
        return pd.DataFrame(columns=headers)

    if dtype == str:
        raw_data = [
            [cell.api.Text for cell in row]
            for row in data_range.rows
        ]
    else:
        raw_data = data_range.value
        if not isinstance(raw_data, (list, tuple)):
            raw_data = [raw_data]
        if not isinstance(raw_data[0], (list, tuple)):
            raw_data = [raw_data]

    df = pd.DataFrame(raw_data, columns=headers)

    # Apply dtype conversions
    if dtype is not None and dtype != str:
        df = df.astype(dtype)
    elif isinstance(dtype, dict):
        for col, col_dtype in dtype.items():
            df[col] = df[col].astype(col_dtype)

    # Strip whitespace from string values if enabled
    if strip:
        df = df.map(lambda x: x.strip() if isinstance(x, str) else x)

    if dropnan:
        df = df.dropna(how='all')

    return df


def read_excel_range(path, sheet_name, cell_range, dtype=None, use_header=True):
    """
    Read a specific Excel range or cell from an Excel file.

    Parameters:
        path (str or Path): Full path to the Excel workbook.
        sheet_name (str): The name of the sheet containing the range.
        cell_range (str): Excel reference (e.g., "B14:L30", "F", "F10").
        dtype (dict or type, optional): Data type(s) to apply to the DataFrame.
        use_header (bool): If True, use the first row of the range as column headers.

    Returns:
        pd.DataFrame or single value: Depending on input range.
    """
    path = Path(path)
    app = xw.App(visible=False)
    try:
        wb = app.books.open(str(path))
        sheet = wb.sheets[sheet_name]

        # Case 1: Single cell like "F10"
        if re.fullmatch(r"[A-Z]+[0-9]+", cell_range, re.IGNORECASE):
            value = sheet.range(cell_range).value
            return value

        # Case 2: Single column like "F"
        elif re.fullmatch(r"[A-Z]+", cell_range, re.IGNORECASE):
            col = cell_range.upper()
            last_row = sheet.range(f"{col}1048576").end("up").row  # Find last used row in column
            cell_range = f"{col}1:{col}{last_row}"

        # Case 3: Range like "B14:L30"
        if use_header:
            data = sheet.range(cell_range).options(pd.DataFrame, header=1, index=False).value
        else:
            values = sheet.range(cell_range).value
            if values is None:
                data = pd.DataFrame()
            else:
                data = pd.DataFrame(values)
                data.columns = [f"Column{i + 1}" for i in range(data.shape[1])]

        if dtype is not None and not data.empty:
            data = data.astype(dtype)
    finally:
        wb.close()
        app.quit()
    return data


def read_static_excel_range(path, sheet_name, cell_range, dtype=None, use_header=True):
    """
    Read a specific Excel range or cell from an Excel file without launching Excel.

    Parameters:
        path (str or Path): Full path to the Excel workbook (.xlsx).
        sheet_name (str): The name of the sheet containing the range.
        cell_range (str): Excel reference (e.g., "B14:L30", "F", "F10").
        dtype (dict or type, optional): Data type(s) to apply to the DataFrame.
        use_header (bool): If True, use the first row of the range as column headers.

    Returns:
        pd.DataFrame, scalar, or None: Depending on input range.
    """
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found in workbook.")
    ws = wb[sheet_name]

    # Case 1: Single cell like "F10"
    if re.fullmatch(r"[A-Z]+[0-9]+", cell_range, re.IGNORECASE):
        return ws[cell_range].value

    # Case 2: Single column like "F"
    elif re.fullmatch(r"[A-Z]+", cell_range, re.IGNORECASE):
        col_letter = cell_range.upper()
        values = [cell.value for cell in ws[col_letter] if cell.value is not None]
        return pd.DataFrame(values, columns=[col_letter])

    # Case 3: Explicit range like "B14:L30"
    else:
        data = ws[cell_range]
        rows = [[cell.value for cell in row] for row in data]

        if not rows:
            return pd.DataFrame()

        if use_header:
            header = rows[0]
            df = pd.DataFrame(rows[1:], columns=header)
        else:
            df = pd.DataFrame(rows)
            df.columns = [f"Column{i+1}" for i in range(df.shape[1])]

        if dtype is not None and not df.empty:
            df = df.astype(dtype)

        return df


def clear_excel_table_contents(workbook_name, sheet_name, table_name):
    """
    Clears the contents (body only) of an Excel Table without deleting the header or table structure.

    Parameters:
        workbook_name (str): The name of the workbook.
        sheet_name (str): The name of the sheet containing the table.
        table_name (str): The name of the Excel Table (not the range name).
    """
    wb = xw.Book(workbook_name)
    sheet = wb.sheets[sheet_name]
    table = sheet.tables[table_name]

    # Clear the contents of the data body range only (not headers or total rows)
    if table.data_body_range:
        table.data_body_range.clear_contents()


def add_unique_row(df1, df2, exclude_columns=None):
    """
    Adds a row from `df1` to `df2` if no equivalent row already exists,
    with optional exclusion of specific columns from the comparison.

    Two rows are considered equivalent if all non-excluded column values match,
    treating NaN and None as equal and attempting to convert numerical values to float
    before comparison.

    Parameters
    ----------
    df1 : pandas.DataFrame
        A single-row DataFrame representing the row to add.
    df2 : pandas.DataFrame
        The target DataFrame to which the row may be added.
    exclude_columns : list of str, optional
        List of column names to ignore when checking for matching rows.
        Defaults to an empty list.

    Returns
    -------
    df2 : pandas.DataFrame
        The updated DataFrame after adding the row if it was unique.
    matching_indices : list of int
        List of indices in `df2` where matching rows were found.
        Empty if the row was added.

    Notes
    -----
    - Missing values (NaN, None) are treated as equivalent during comparison.
    - If the columns of `df1` and `df2` differ, missing columns are filled with None.
    - All values are cast to string after float conversion to ensure robust comparison.
    """
    if exclude_columns is None:
        exclude_columns = []

    all_columns = df1.columns.union(df2.columns)
    df1 = df1.reindex(columns=all_columns, fill_value=None)
    df2 = df2.reindex(columns=all_columns, fill_value=None)

    # Drop excluded columns from both df1 and df2 for comparison
    df1_comp = df1.drop(columns=exclude_columns, errors='ignore')
    df2_comp = df2.drop(columns=exclude_columns, errors='ignore')

    # Convert both DataFrames to treat NaN and None as equal and convert numerical values (including strings) to float
    def convert_to_float(x):
        try:
            return float(x)
        except (ValueError, TypeError):
            return np.nan if pd.isna(x) else x

    df1_comp = df1_comp.map(convert_to_float)
    df2_comp = df2_comp.map(convert_to_float)

    # Ensure all columns are of the same type for proper comparison
    df1_comp = df1_comp.astype(str)
    df2_comp = df2_comp.astype(str)

    # Check for rows in df2 that match the row in df1
    matching_rows = df2_comp[df2_comp.eq(df1_comp.values[0]).all(axis=1)]

    if not matching_rows.empty:
        matching_indices = matching_rows.index.tolist()
    else:
        df2 = pd.concat([df2, df1], ignore_index=True)
        matching_indices = []

    return df2, matching_indices


def call_vba_dropdown_macro(workbook_name: str, sheet_name: str, dropdown_name: str, new_value: str):
    wb = xw.Book(workbook_name)  # Adjust path or use xw.Book.caller()
    wb.macro('set_dropdown_value')(sheet_name, dropdown_name, new_value)


def insert_plot(fig, workbook_name, sheet_name, named_range):
    """
    Insert a Matplotlib Figure into an already open Excel workbook at the named range.

    Parameters:
    - fig: matplotlib.figure.Figure object
    - workbook_name: str, name of the open Excel workbook (e.g., 'file.xlsx')
    - sheet_name: str, name of the sheet in the workbook
    - named_range: str, named range in the sheet to place the image at
    """

    app = xw.apps.active
    wb = app.books[workbook_name]
    sheet = wb.sheets[sheet_name]
    rng = sheet.range(named_range)

    with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmpfile:
        fig.savefig(tmpfile.name, bbox_inches='tight')
        tmpfile.flush()

        sheet.pictures.add(tmpfile.name,
                           name=f"Fig_{named_range}",
                           update=True,
                           top=rng.top,
                           left=rng.left)
    os.remove(tmpfile.name)


def read_named_range(path, name, sheet_name=None, dtype=None, use_header=True):
    """
    Read a named range from an Excel file (supports workbook-level and sheet-level names).
    """
    path = Path(path)
    app = xw.App(visible=False)
    try:
        wb = app.books.open(str(path))

        rng = None

        # 1) Workbook-level names
        for n in wb.names:
            if n.name == name or n.name.endswith(f"!{name}"):
                rng = n.refers_to_range
                break

        # 2) Sheet-level names
        if rng is None and sheet_name is not None:
            for n in wb.sheets[sheet_name].names:
                if n.name == name or n.name.endswith(f"!{name}"):
                    rng = n.refers_to_range
                    break

        if rng is None:
            raise KeyError(
                f"Named range '{name}' not found in workbook or sheet '{sheet_name}'."
            )

        # Single cell → scalar
        if rng.rows.count == 1 and rng.columns.count == 1:
            return rng.value

        # Multi-cell → DataFrame
        if use_header:
            data = rng.options(pd.DataFrame, header=1, index=False).value
        else:
            values = rng.value
            if values is None:
                data = pd.DataFrame()
            else:
                data = pd.DataFrame(values)
                data.columns = [f"Column{i + 1}" for i in range(data.shape[1])]

        if dtype is not None and not data.empty:
            data = data.astype(dtype)

    finally:
        wb.close()
        app.quit()

    return data