# %% helpers
import os
import os.path
import re

import numpy as np
import pandas as pd
from pandas.api.types import CategoricalDtype

import excel as ex
import misc as mc

os.environ["MPLBACKEND"] = "Agg"  # kein GUI nötig

# optional zusätzlich:
try:
    import matplotlib

    matplotlib.use("Agg", force=True)
except Exception:
    pass


def add_node(df, z_new, defaults=None, add_outside_bound=False):
    """
    Adds a new node at the specified height `z_new` to the node DataFrame `df`.

    If `z_new` already exists in the 'Elevation [m]' column, or lies outside the vertical bounds
    and `add_outside_bound` is False, the original DataFrame is returned unchanged.

    The new node includes default values for all required fields. For any additional columns
    in `df`, values are taken from the `defaults` dictionary if provided by type (e.g., 'float', 'int').
    If a dtype is not in `defaults`, a general type-based fallback is assigned.

    Parameters
    ----------
    df : pd.DataFrame
        Original node DataFrame. Must include at least:
            - 'Elevation [m]'
            - 'node'
            - 'pInertia'
            - 'pMass'

    z_new : float
        Elevation at which to insert the new node.

    defaults : dict, optional
        Optional dictionary mapping data type names (e.g., 'float', 'bool') to default values.

    add_outside_bound : bool, default False
        If False, prevents insertion of nodes outside the current elevation range.

    Returns
    -------
    pd.DataFrame
        Updated DataFrame with the new node added, sorted by elevation and reindexed.
    """
    if z_new in df['Elevation [m]'].values:
        return df

    if not add_outside_bound:
        z_min, z_max = df['Elevation [m]'].min(), df['Elevation [m]'].max()
        if not (z_min < z_new < z_max):
            return df

    defaults = defaults or {}

    new_row = {
        'Elevation [m]': z_new,
        'node': 0,  # placeholder
    }

    for col in df.columns:
        if col in new_row:
            continue

        dtype = df[col].dtype

        if pd.api.types.is_bool_dtype(dtype):
            new_row[col] = defaults.get("bool", False)
        elif pd.api.types.is_integer_dtype(dtype):
            new_row[col] = defaults.get("int", np.nan)
        elif pd.api.types.is_float_dtype(dtype):
            new_row[col] = defaults.get("float", np.nan)
        elif pd.api.types.is_numeric_dtype(dtype):  # fallback
            new_row[col] = defaults.get("numeric", np.nan)
        elif pd.api.types.is_datetime64_any_dtype(dtype):
            new_row[col] = defaults.get("datetime", pd.NaT)
        elif pd.api.types.is_object_dtype(dtype):
            new_row[col] = defaults.get("object", None)
        else:
            new_row[col] = defaults.get("other", None)

    df_updated = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df_updated = df_updated.sort_values(by='Elevation [m]', ascending=False).reset_index(drop=True)

    max_node = df_updated['node'].max() + 1
    df_updated['node'] = list(range(max_node, max_node - len(df_updated), -1))

    return df_updated


def check_values(df: pd.DataFrame, columns=None, mode='missing') -> list[str]:
    """
    Check for missing or present values in specified columns of a DataFrame.

    Parameters:
    - df (pd.DataFrame): The DataFrame to check.
    - columns (list, optional): List of column names to check. Defaults to all columns.
    - mode (str): 'missing' to check for NaN/None values, 'present' to check for any present values.

    Returns:
    - list[str]: A list of error message strings describing the found issues.
    """
    if columns is None:
        columns = df.columns
    else:
        missing_cols = [col for col in columns if col not in df.columns]
        if missing_cols:
            raise ValueError(f"The following columns are not in the DataFrame: {missing_cols}")

    error_messages = []

    for col in columns:
        if mode == 'missing':
            mask = df[col].isna()
            if mask.any():
                for idx in df.index[mask]:
                    error_messages.append(f"Missing value in column '{col}' at index {idx}")
        elif mode == 'present':
            mask = ~df[col].isna()
            if mask.any():
                for idx in df.index[mask]:
                    error_messages.append(f"Unexpected value in column '{col}' at index {idx}: {df.at[idx, col]!r}")
        else:
            raise ValueError("Invalid mode. Use 'missing' or 'present'.")

    return error_messages


def interpolate_with_neighbors(x, y):
    """
    Fills holes (None or NaN) in y using linear interpolation:
    - Interior holes use left and right known neighbors.
    - Leading holes use the first two known values on the right.
    - Trailing holes use the last two known values on the left.

    Parameters:
    - x: list of x-values (same length as y), must be sorted
    - y: list of y-values with possible holes (None or np.nan)

    Returns:
    - A new list with the holes in y filled locally
    """
    x = list(x)
    y = list(y)
    y_filled = y.copy()
    n = len(y)

    i = 0
    while i < n:
        if y_filled[i] is None or (isinstance(y_filled[i], float) and np.isnan(y_filled[i])):
            # Start of hole
            start = i
            while i < n and (y_filled[i] is None or (isinstance(y_filled[i], float) and np.isnan(y_filled[i]))):
                i += 1
            end = i  # first known value after the hole

            # Case 1: Interior hole
            if start > 0 and end < n:
                x0, y0 = x[start - 1], y_filled[start - 1]
                x1, y1 = x[end], y_filled[end]
                for j in range(start, end):
                    t = (x[j] - x0) / (x1 - x0)
                    y_filled[j] = (1 - t) * y0 + t * y1

            # Case 2: Leading edge hole
            elif end < n:
                # Use first two known values after the hole
                k1 = end
                while k1 + 1 < n and (y_filled[k1 + 1] is None or np.isnan(y_filled[k1 + 1])):
                    k1 += 1
                if k1 + 1 < n:
                    x0, y0 = x[k1], y_filled[k1]
                    x1, y1 = x[k1 + 1], y_filled[k1 + 1]
                    for j in range(start, end):
                        t = (x[j] - x0) / (x1 - x0)
                        y_filled[j] = (1 - t) * y0 + t * y1

            # Case 3: Trailing edge hole
            elif start > 1:
                # Use last two known values before the hole
                k1 = start - 1
                while k1 - 1 >= 0 and (y_filled[k1 - 1] is None or np.isnan(y_filled[k1 - 1])):
                    k1 -= 1
                if k1 - 1 >= 0:
                    x0, y0 = x[k1 - 1], y_filled[k1 - 1]
                    x1, y1 = x[k1], y_filled[k1]
                    for j in range(start, end):
                        t = (x[j] - x0) / (x1 - x0)
                        y_filled[j] = (1 - t) * y0 + t * y1

            # Otherwise: cannot fill
        else:
            i += 1

    return y_filled


def read_lua_values(file_path, keys):
    """
    Extracts specified key-value pairs from a Lua file.

    Parameters:
    ----------
    file_path : str
        The path to the Lua file from which values need to be extracted.

    keys : list of str
        A list of keys (as strings) for which the corresponding values should be extracted from the Lua file.

    Returns:
    -------
    dict
        A dictionary where the keys are the specified keys from the input list, and the values are the corresponding values
        found in the Lua file. The values are converted to their appropriate types (int, float, bool, or str) based on their
        format in the Lua file.
    """

    # Dictionary to store the values
    values_dict = {}

    # Regular expression to match key-value pairs in the Lua file
    pattern = re.compile(r'(\w+)\s*=\s*(.+)')

    with open(file_path, 'r') as file:
        for line in file:
            # Remove comments and strip any leading/trailing whitespace
            line = line.split('--')[0].strip()
            if not line:
                continue

            # Match the key-value pair
            match = pattern.match(line)
            if match:
                key, value = match.groups()

                # Remove trailing comma if present
                value = value.rstrip(',')

                # Check if the key is in the desired keys
                if key in keys:
                    # Attempt to convert to a number or leave as a string
                    try:
                        # Handle numbers and booleans
                        if value.lower() == "true":
                            value = True
                        elif value.lower() == "false":
                            value = False
                        elif "." in value:
                            value = float(value)
                        else:
                            value = int(value)
                    except ValueError:
                        # Keep as string if conversion fails
                        value = value.strip('"').strip("'")

                    values_dict[key] = value

    return values_dict


def write_lua_variables(lua_string, variables):
    """
    Updates Lua variable assignments in a string.

    Args:
        lua_string (str): The Lua source code as a string.
        variables (dict): Dictionary of variables to update, e.g., {'var_name': new_value}.

    Returns:
        str: The modified Lua code as a string.
    """
    lines = lua_string.splitlines()

    for variable_name, new_value in variables.items():
        variable_pattern = re.compile(rf'^(\s*{re.escape(variable_name)}\s*=\s*).*?(\s*,\s*--.*)?$')

        for i, line in enumerate(lines):
            match = variable_pattern.match(line)
            if match:
                indentation = match.group(1)
                rest_of_line = match.group(2) if match.group(2) else ','
                lines[i] = f"{indentation}{new_value}{rest_of_line}"
                break

    return "\n".join(lines)


# %% functions

from typing import Tuple


def calculate_deflection(
        NODES: pd.DataFrame,
        defl_MP: Tuple[float, str],
        defl_TP: Tuple[float, str],
        defl_Tower: Tuple[float, str],
) -> pd.Series:
    """
    Calculate continuous deflection values with varying tilt angles.
    - defl_Tower is RELATIVE to defl_TP
    - defl_MP and defl_TP are ABSOLUTE
    The deflection at the lowest elevation is always zero, including BORDER nodes.

    Parameters:
    - NODES: DataFrame with ["Elevation [m]", "Affiliation"]
    - defl_MP, defl_TP, defl_Tower: Tuples (value, unit)

    Returns:
    - pd.Series of deflection values
    """

    def _convert_to_rad(value: float, unit: str) -> float:
        if unit == "deg":
            return np.deg2rad(value)
        elif unit == "mm/m":
            return np.arctan(value / 1000)
        else:
            raise ValueError(f"Unsupported unit '{unit}'")

    def _segment_deflection(z: pd.Series, angle_rad: float, base_z: float, offset: float) -> pd.Series:
        return np.tan(angle_rad) * (z - base_z) + offset

    # Extract and sort elevations
    z = NODES["Elevation [m]"]
    affiliations = NODES["Affiliation"]

    # Sort BORDER elevations to determine boundaries
    borders = NODES.loc[affiliations == "BORDER", "Elevation [m]"].sort_values().values
    if len(borders) < 2:
        raise ValueError("Two 'BORDER' nodes required to separate TP and TOWER.")

    base_TOWER = borders[1]
    base_TP = borders[0]
    base_MP = z.min()  # Always take the lowest z as the MP base

    # Convert angles
    angle_MP = _convert_to_rad(*defl_MP)
    angle_TP = _convert_to_rad(*defl_TP)
    angle_Tower = angle_TP + _convert_to_rad(*defl_Tower)  # relative to TP

    # Compute deflection offsets for continuity
    defl_MP_end = np.tan(angle_MP) * (base_TP - base_MP)
    defl_TP_end = np.tan(angle_TP) * (base_TOWER - base_TP) + defl_MP_end

    # Prepare full output Series
    defl = pd.Series(0.0, index=NODES.index, dtype="float64")

    # MP and BORDER nodes below TP
    mask_MP = (z <= base_TP)
    defl[mask_MP] = _segment_deflection(z[mask_MP], angle_MP, base_MP, 0)

    # TP and BORDER nodes between TP and TOWER
    mask_TP = (z > base_TP) & (z <= base_TOWER)
    defl[mask_TP] = _segment_deflection(z[mask_TP], angle_TP, base_TP, defl_MP_end)

    # TOWER and BORDER nodes above TOWER
    mask_TOWER = z > base_TOWER
    defl[mask_TOWER] = _segment_deflection(z[mask_TOWER], angle_Tower, base_TOWER, defl_TP_end)

    return defl


def create_JBOOST_struct(GEOMETRY, RNA, defl_MP, delf_TOWER, MASSES=None, MARINE_GROWTH=None, defl_TP=None,
                         ModelName="Struct", EModul="2.10E+11", fyk="355", poisson="0.3", dens="7850", addMass=0,
                         member_id=1, create_node_tolerance=0.1, seabed_level=None, waterlevel=0):
    """
    Generates a JBOOST structural input text block based on geometric and mass data for offshore wind turbine structures.
    """

    NODES_txt = []
    ELEMENTS_txt = []
    elements = []
    waterlevel = float(waterlevel)

    # Filter geometry below seabed
    if seabed_level is not None:
        GEOMETRY = mc.add_element(GEOMETRY, seabed_level)
    GEOMETRY = GEOMETRY.loc[GEOMETRY["Bottom [m]"] >= seabed_level]

    # Extract and initialize nodes
    NODES = mc.extract_nodes_from_elements(GEOMETRY)
    nodes = list(reversed([501 + i for i in range(len(NODES))]))
    NODES["node"] = nodes
    NODES["pInertia"] = 0
    NODES["comment"] = None
    NODES["added"] = False
    # Deflection
    NODES["DEFL"] = calculate_deflection(NODES, defl_MP, defl_TP, delf_TOWER)

    # Insert node at water level
    if waterlevel not in NODES["Elevation [m]"].values:
        NODES = add_node(NODES, waterlevel, defaults={"float": 0})
        GEOMETRY = mc.add_element(GEOMETRY, waterlevel)
        NODES.loc[NODES["Elevation [m]"] == waterlevel, "comment"] = "water level "
        NODES.loc[NODES["Elevation [m]"] == waterlevel, "added"] = True

    # Add marine growth
    if MARINE_GROWTH is not None:

        z_marine = MARINE_GROWTH["Bottom [m]"].to_list() + [MARINE_GROWTH.iloc[1]["Top [m]"]]
        z_marine = [round(z, 1) for z in z_marine]

        for z in z_marine:
            if not any(np.isclose(z, NODES["Elevation [m]"].values)):
                NODES = add_node(NODES, z, defaults={"float": 0})
                GEOMETRY = mc.add_element(GEOMETRY, z)
                NODES.loc[NODES["Elevation [m]"] == z, "comment"] = "marine growth border"
                NODES.loc[NODES["Elevation [m]"] == z, "added"] = True

    NODES["pMass"] = 0.0
    NODES["pMassNames"] = None
    # Add point masses
    if MASSES is not None:
        for idx in MASSES.index:
            z_bot = MASSES.loc[idx, "Bottom [m]"]
            z_Mass = (z_bot + MASSES.loc[idx, "Top [m]"]) / 2 if pd.notna(z_bot) else MASSES.loc[idx, "Top [m]"]

            differences = np.abs(NODES["Elevation [m]"].values - z_Mass)
            within_tol = np.where(differences <= create_node_tolerance)[0]

            # if node is (nearly) on Node
            if len(within_tol) > 0:
                closest_index = within_tol[np.argmin(differences[within_tol])]
                NODES.loc[closest_index, "pMass"] += MASSES.loc[idx, "Mass [kg]"]

                if NODES.loc[closest_index, "pMassNames"] is None:
                    NODES.loc[closest_index, "pMassNames"] = MASSES.loc[idx, "Name"] + " "
                else:
                    NODES.loc[closest_index, "pMassNames"] += MASSES.loc[idx, "Name"] + " "

            # if node mass is over bottom
            elif z_Mass >= GEOMETRY["Bottom [m]"].values[-1]:

                # add Node and Element
                NODES = add_node(NODES, z_Mass, defaults={"float": 0})
                GEOMETRY = mc.add_element(GEOMETRY, z_Mass)
                NODES.loc[NODES["Elevation [m]"] == z_Mass, "added"] = True
                NODES.loc[NODES["Elevation [m]"] == z_Mass, "comment"] = None

                NODES.loc[NODES["Elevation [m]"] == z_Mass, "pMass"] += MASSES.loc[idx, "Mass [kg]"]

                NODES.loc[NODES["Elevation [m]"] == z_Mass, "pMassNames"] = MASSES.loc[idx, "Name"] + " "

            else:
                print(f"Warning! Mass '{MASSES.loc[idx, 'Name']}' not added, it is below the seabed level!")

    # Add RNA
    MP_top = NODES.iloc[0]["Elevation [m]"]
    D_MP_top = GEOMETRY.iloc[0]["D, top [m]"]
    t_MP_top = GEOMETRY.iloc[0]["t [mm]"]

    z_RNA = MP_top + RNA.loc[0, "Vertical Offset TT to HH [m]"]
    NODES = add_node(NODES, z_RNA, defaults={"float": 0}, add_outside_bound=True)
    GEOMETRY = pd.concat([
        pd.DataFrame({
            'Affiliation': 'TOWER',
            'Top [m]': z_RNA,
            'Bottom [m]': MP_top,
            'D, top [m]': D_MP_top,
            'D, bottom [m]': D_MP_top,
            't [mm]': t_MP_top
        }, index=[-1]),
        GEOMETRY
    ], ignore_index=True)

    NODES.loc[NODES["Elevation [m]"] == z_RNA, "pMass"] += RNA.loc[0, "Mass of RNA [kg]"]
    NODES.loc[NODES["Elevation [m]"] == z_RNA, "comment"] = f"RNA {RNA.loc[0, 'Identifier']}"
    NODES.loc[NODES["Elevation [m]"] == z_RNA, "added"] = True
    NODES.loc[NODES["Elevation [m]"] == z_RNA, "pMassNames"] = None

    NODES.loc[NODES["Elevation [m]"] == z_RNA, "pInertia"] = (
                                                                     RNA.loc[
                                                                         0, 'Inertia of RNA fore-aft @COG [kg m^2]'] +
                                                                     RNA.loc[
                                                                         0, 'Inertia of RNA side-side @COG [kg m^2]']
                                                             ) / 2

    # Interpolate deflections for added nodes and reverse node order
    NODES["DEFL"] = interpolate_with_neighbors(NODES["Elevation [m]"].values, NODES["DEFL"].values)
    NODES = NODES.iloc[::-1].reset_index(drop=True)

    # Create node definitions
    for _, node in NODES.iterrows():
        elevation = round(node['Elevation [m]'], 2)
        pMass = round(node['pMass'], 0)
        pInertia = node['pInertia']  # keep as-is, assuming no formatting needed
        pDefl = round(node['DEFL'], 3)

        line = (
            f"os_FeNode{{model=ModelName"
            f"\t,node={int(node['node'])}"
            f"\t,x=0\t,y=0"
            f"\t,z={elevation:.2f}"
            f"\t,pMass={pMass:06.0f}"
            f"\t,pInertia={pInertia}"
            f"\t,pOutOfVertically={pDefl:06.3f}}}"
        )
        comment = ""
        if node["added"]:
            comment += "-- added node,  "
            if node["comment"]:
                comment += node["comment"] + " "
            if node["pMassNames"]:
                comment += "Mass(es): " + node["pMassNames"]

        else:
            if node["pMassNames"]:
                comment += "-- added masses on existing node: " + node["pMassNames"]
        line += comment

        NODES_txt.append(line)

    # Create element definitions
    for i in range(len(NODES) - 1):
        startnode = int(NODES.loc[i, "node"])
        endnode = int(NODES.loc[i + 1, "node"])
        diameter = round((GEOMETRY.loc[i, "D, top [m]"] + GEOMETRY.loc[i, "D, bottom [m]"]) / 2, 2)
        t_wall = round(GEOMETRY.loc[i, "t [mm]"] / 1000, 3)

        elements.append({
            "elem_id": i + 1,
            "startnode": startnode,
            "endnode": endnode,
            "diameter": diameter,
            "t_wall": t_wall,
            "dens": dens
        })

    ELEM = pd.DataFrame(elements)
    ELEM.at[ELEM.index[-1], "dens"] = 1.0

    for _, elem in ELEM.iterrows():
        elem_id = int(elem['elem_id'])
        startnode = int(elem['startnode'])
        endnode = int(elem['endnode'])
        diameter = round(elem['diameter'], 2)
        t_wall = round(elem['t_wall'], 3)
        dens_rounded = round(elem['dens'], 3)

        line = (
            f"os_FeElem{{model=ModelName"
            f"\t,elem_id={elem_id:03.0f}"
            f"\t,startnode={startnode:03.0f}"
            f"\t,endnode={endnode:03.0f}"
            f"\t,diameter={diameter:.2f}"
            f"\t,t_wall={t_wall:.3f}"
            f"\t,EModul={EModul}"
            f"\t,fky={fyk}"
            f"\t,poisson={poisson}"
            f"\t,dens={dens_rounded}"
            f"\t,addMass={addMass}"
            f"\t,member_id={member_id}}}"
        )
        ELEMENTS_txt.append(line)

    # Compose final JBOOST input
    text = (
            "--Input for JBOOST generated by Excel\n"
            "--    Definition Modelname\n"
            f'local	ModelName="{ModelName}"\n'
            "--    Definition der FE-Knoten\n"
            + "\n".join(NODES_txt) + "\n\n"
                                     "--Definition der FE-Elemente- Zusatzmassen in kg / m\n"
            + "\n".join(ELEMENTS_txt) + "\n"
    )

    return text


def create_JBOOST_proj(Parameters, marine_growth=None, modelname="struct.lua", runFEModul=True, runFrequencyModul=False,
                       runHindcastValidation=False, wavefile="wave.lua",
                       windfile="wind."):
    """
       Generates a Lua-based project file text for JBOOST simulations based on input parameters and configuration flags.

       The function creates a Lua script for structural simulation using JBOOST. It allows configuration of modules such
       as FEM, frequency domain, and hindcast validation. The function also embeds parameters and optional marine growth
       data into the Lua template.

       Parameters
       ----------
       Parameters : dict
           A dictionary containing configuration values to be injected into the Lua script using placeholder tags.
       marine_growth : pandas.DataFrame, optional
           DataFrame with marine growth layer specifications. Expected columns:
           - 'Top [m]': Top elevation of the layer.
           - 'Bottom [m]': Bottom elevation of the layer.
           - 'Marine Growth [mm]': Thickness of the marine growth layer in millimeters.
       modelname : str, default="struct.lua"
           Filename for the structural model Lua file.
       runFEModul : bool, default=True
           If True, includes a call to run the FEM module (`os_RunFEModul()`).
       runFrequencyModul : bool, default=False
           If True, includes a call to run the frequency domain module (`os_RunFrequencyDomain()`).
       runHindcastValidation : bool, default=False
           If True, includes a call to run hindcast validation (`os_RunHindcastValidation()`).
       wavefile : str, default="wave.lua"
           Filename for the wave input Lua file.
       windfile : str, default="wind."
           Filename for the wind input Lua file.

       Returns
       -------
       str
           A string containing the fully-formed Lua project script for use with JBOOST.

       Raises
       ------
       ValueError
           If both `runFEModul` and `runFrequencyModul` are set to True. Only one simulation mode should be active.

       Notes
       -----
       - Placeholder tokens in the Lua template (e.g., `?modelname`, `?RunFeModul`, etc.) are replaced based on inputs.
       - Marine growth layers, if provided, are inserted via `os_OceanMG{}` calls.
       - Assumes `write_lua_variables()` is defined elsewhere to handle parameter injection.
       """
    if runFEModul and runFrequencyModul:
        raise ValueError("RunFEModul and runFrequencyModul can not both be set to True.")
    Proj_text = """\
-- ++++++++++++++++++++++++++++++++++++++++ Model Data +++++++++++++++++++++++++++++++++++++++++++++++++++

model = getdata(?modelname)
model()
wave = (getdata(?wavefile))
wave()
wind = getdata(?wavefile)
wind()

-- ++++++++++++++++++++++++++++++++++++++++ Configurations +++++++++++++++++++++++++++++++++++++++++++++++++++

os_LoadConfigGeneral(
{
    Config_name       = ,
    Result_name       = ,
    Scatter_name      = ,
    FeModel_name      = ,
    Wind_name         = ,
    Ocean_name        = "test",
}
)

-- ++++++++++++++++++++++++++++++++++++++++ FEModul Konfiguration ++++++++++++++++++++++++++++++++++++++++++++++++++++++
os_LoadConfigFEModul(
{
    foundation_superelement  = ,  -- Knotennummer (-1=keine Ersatzmatizen fuer Pfahl, >0 Ersatzsteifigkeit bzw. -massen)

    found_stiff_trans        = ,  -- aus 22A559 EnBW GOA aber 10.4 m pile
    found_stiff_rotat        = ,
    found_stiff_coupl        = ,
    found_mass_trans         = ,
    found_mass_rotat         = ,
    found_mass_coupl         = ,

    shearCorr                = ,  -- Schubkorrekturfaktor (2.0 fuer Kreisquerschnitt und Timoshenko 0.0 fuer Bernoulli)
    res_NumEF                = ,  -- Anzahl der Eigenfrequenzen, die ausgegeben werden
    hydro_add_mass           = ,  -- Added Mass fuer analysen der Strukturdynamik 1 = mit added mass
}
)

?MARINE_GROWTH

os_LoadConfigOcean(
{
    water_density    = ,  -- water density in [kg/m ]
    water_level      = ,  -- wrt LAT in [m] !!! there must be a node @ this height !!!
    seabed_level     = ,  -- wrt LAT in [m]
    growth_density   = ,  -- marine growth density in [kg/m ]
}
)

os_LoadConfigFrequModul(
{
    frequRange            = ,  -- 0 Hz to frequRange in Hz
    frequRangeSolution    = ,  -- no. of frequency components
    maxEFcalc             = ,  -- max number of considered frequencies
    damping_struct        = ,  -- damping ratio structural (material, viscous, soil)
    damping_tower         = ,
    damping_aerodyn       = ,  -- damping ratio aerodynamic (weighted mean value over wind speeds)
    design_life           = ,  -- structural design life in years
    N_ref                 = ,  -- no. of cycles for DEL calculation
    tech_availability     = ,  -- technical availability turbine in percent --> 0.95*30y/31.0y
    -- Note that this coers also the commissioning and decommissioning times
    SN_slope              = ,  -- S-N slope material
    h_refwindspeed        = ,  -- height reference in m of wind speed data in wind wave scatter
    h_hub                 = ,  -- hub height wrt LAT in m
    height_exp            = ,  -- height exponent
    -- height exp was set specifically here so that the average wind speed from the ROUGH Vw Hs scatter is 10.5 m/s at hub height
    TM02_period           = ,  -- [0] if peak periods stated, [1] if zero crossing Tm02 periods stated -> noch nicht eingebaut...FOs
    refineScatter         = ,  -- refinement of scatter data on Tp axis (factor 10 coded)  [0]=off; [1]=on; [2]=on incl. validation plots!
    fullScatter           = ,  -- calc full real scatter FLS for validation purpose
    WindspecDataBase      = ,  -- path on windspec database (*.csv) or "-" in case no wind load is to be considered
    res_Nodes             = 
}
)
-- ----------------------------------------------
-- Execute program steps
-- ----------------------------------------------

?RunFeModul
?RunFrequencyDomain
?runHindcastValidation

-- Ausgabe der Textdateien
os_WriteResultsText([[Results_JBOOST_Text]])
os_PlotResultsGraph([[Result_JBOOST_Graph]])
"""

    Proj_text = write_lua_variables(Proj_text, Parameters)

    if runFEModul:
        Proj_text = Proj_text.replace("?RunFeModul", "os_RunFEModul()")
    else:
        Proj_text = Proj_text.replace("?RunFeModul", "")

    if runFrequencyModul:
        Proj_text = Proj_text.replace("?RunFrequencyDomain", "os_RunFrequencyDomain()")
    else:
        Proj_text = Proj_text.replace("?RunFrequencyDomain", "")

    if runHindcastValidation:
        Proj_text = Proj_text.replace("?runHindcastValidation", "os_RunHindcastValidation()")
    else:
        Proj_text = Proj_text.replace("?runHindcastValidation", "")

    Proj_text = Proj_text.replace("?modelname", modelname)
    Proj_text = Proj_text.replace("?windfile", windfile)
    Proj_text = Proj_text.replace("?wavefile", wavefile)

    # marine growth
    if marine_growth is not None:
        marine_text = [
            "os_OceanMG{" + f"id=HD, topMGsection = {row['Top [m]']},   bottomMGsection = {row['Bottom [m]']},    t={row['Marine Growth [mm]'] / 1000}" + "} -- marine growth layer spec."
            for _, row in marine_growth.iterrows()]
        marine_text = "\n".join(marine_text)
        Proj_text = Proj_text.replace("?MARINE_GROWTH", marine_text)

    return Proj_text


def create_WLGen_file(APPURTANCES, ADDITIONAL_MASSES, MP, TP, MARINE_GROWTH, skirt=None):
    """
    Generate a WLGen input file text based on structural and geometric input data.

    Parameters
    ----------
    APPURTANCES : pd.DataFrame
        Table containing appurtenance data. Required columns:
        'Top [m]', 'Bottom [m]', 'Mass [kg]', 'Diameter [m]', 'Orientation [°]',
        'Surface roughness [m]', 'Distance Axis to Axis [m]', 'Gap between surfaces [m]', 'Name'.
    ADDITIONAL_MASSES : pd.DataFrame
        Table of additional point masses. Required columns:
        'Top [m]', 'Bottom [m]', 'Mass [kg]', 'Name'.
    MP : pd.DataFrame
        Monopile section data. Required columns:
        'D, bottom [m]', 'D, top [m]', 't [mm]', 'Bottom [m]', 'Top [m]'.
    TP : pd.DataFrame
        Transition piece section data. Same format as MP.
    MARINE_GROWTH : pd.DataFrame
        Marine growth parameters. Required columns:
        'Bottom [m]', 'Top [m]', 'Marine Growth [mm]', 'Density  [kg/m^3]', 'Surface Roughness [m]'.
    skirt: pd.DataFrame, optional

    Returns
    -------
    tuple[str, str]
        A tuple where the first item is the generated WLGen text if successful,
        and the second item is an error message (empty string if successful).
        If an error occurs, the first item is False.
    """

    def check_values(df, columns):
        return [col for col in columns if df[col].isnull().any()]

    input_checks = [
        # For APPURTANCES: exclude mutually exclusive fields from missing value check
        (APPURTANCES, [
            "Top [m]", "Bottom [m]", "Mass [kg]",
            "Diameter [m]", "Orientation [°]", "Surface roughness [m]", "Name"
        ], "APPURTANCES"),
        (ADDITIONAL_MASSES, ["Top [m]", "Bottom [m]", "Mass [kg]", "Name"], "ADDITIONAL_MASSES"),
        (MP, ["D, bottom [m]", "D, top [m]", "t [mm]", "Bottom [m]", "Top [m]"], "MP"),
        (TP, ["D, bottom [m]", "D, top [m]", "t [mm]", "Bottom [m]", "Top [m]"], "TP"),
        (MARINE_GROWTH, [
            "Bottom [m]", "Top [m]", "Marine Growth [mm]",
            "Density  [kg/m^3]", "Surface Roughness [m]"
        ], "MARINE_GROWTH"),
    ]

    for df, required_cols, name in input_checks:
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            return False, f"Missing required columns in {name}:\n" + "\n".join(missing_cols)

        missing_vals = check_values(df, required_cols)
        if missing_vals:
            return False, f"Missing values in {name}:\n" + "\n".join(missing_vals)

    # Additional check: ensure APPURTANCES has the mutually exclusive columns
    for col in ["Distance Axis to Axis [m]", "Gap between surfaces [m]"]:
        if col not in APPURTANCES.columns:
            return False, f"Missing required column '{col}' in APPURTANCES."

    APPURTANCES = APPURTANCES.sort_values(by='Top [m]', ascending=False)
    ADDITIONAL_MASSES = ADDITIONAL_MASSES.sort_values(by='Top [m]', ascending=False)

    # Now validate mutual exclusivity per row
    err_list = []
    for idx, row in APPURTANCES.iterrows():
        row_num = idx + 1
        axis_to_axis = row["Distance Axis to Axis [m]"]
        gap_between = row["Gap between surfaces [m]"]

        if pd.isna(axis_to_axis) and pd.isna(gap_between):
            err_list.append(f"Row {row_num}: Define either 'Distance Axis to Axis [m]' or 'Gap between surfaces [m]'.")
        elif not pd.isna(axis_to_axis) and not pd.isna(gap_between):
            err_list.append(
                f"Row {row_num}: Define only one of 'Distance Axis to Axis [m]' or 'Gap between surfaces [m]', not both.")

    if err_list:
        return False, "Geometry specification issues in APPURTANCES:\n" + "\n".join(err_list)

    # Add skirt
    if skirt is not None:
        skirt_nodes = list(skirt.loc[:, "Top [m]"].values) + [skirt.iloc[-1, :].loc["Bottom [m]"]]
        for skirt_node in skirt_nodes:
            MP = mc.add_element(MP, skirt_node)
            TP = mc.add_element(TP, skirt_node)
        # interpolate skirt nodes
        MP_nodes = list(MP.loc[:, "Top [m]"].values) + [MP.iloc[-1, :].loc["Bottom [m]"]]
        TP_nodes = list(TP.loc[:, "Top [m]"].values) + [TP.iloc[-1, :].loc["Bottom [m]"]]
        all_nodes = TP_nodes[0:-1] + MP_nodes
        overlaps = [node for node in all_nodes if node < float(skirt_nodes[0]) and node > float(skirt_nodes[1])]

        for overlap in overlaps:
            skirt = mc.add_element(skirt, float(overlap))
        for idx, row in skirt.iterrows():
            top = row["Top [m]"]
            bottom = row["Bottom [m]"]

            if top in list(MP.loc[:, "Top [m]"].values):
                MP.loc[MP["Top [m]"] == top, "D, top [m]"] = row["D, top [m]"]
                MP.loc[MP["Bottom [m]"] == bottom, "D, bottom [m]"] = row["D, bottom [m]"]

            if top in list(TP.loc[:, "Top [m]"].values):
                TP.loc[TP["Top [m]"] == top, "D, top [m]"] = row["D, top [m]"]
                TP.loc[TP["Bottom [m]"] == bottom, "D, bottom [m]"] = row["D, bottom [m]"]

    # === STRING GENERATION ===
    Data_MonopileSections = [
        ("Data_MonopileSections{" +
         f" diameter_bot =  {row['D, bottom [m]']:01.3f}, " +
         f" diameter_top = {row['D, top [m]']:01.3f}, " +
         f" wall_thickness = {(row['t [mm]'] / 1000):01.3f}, " +
         f" z_bot = {row['Bottom [m]']:01.3f}, " +
         f" z_top = {row['Top [m]']:01.3f}, " +
         f" surface_roughness = {0:.3f}" +
         "}") for _, row in MP.iterrows()
    ]

    Data_TransitionPieceSections = [
        ("Data_TransitionPieceSections{" +
         f" diameter_bot =  {row['D, bottom [m]']:01.3f}, " +
         f" diameter_top = {row['D, top [m]']:01.3f}, " +
         f" wall_thickness = {(row['t [mm]'] / 1000):01.3f}, " +
         f" z_bot = {row['Bottom [m]']:01.3f}, " +
         f" z_top = {row['Top [m]']:01.3f}, " +
         f" surface_roughness = {0:.3f}" +
         "}") for _, row in TP.iterrows()
    ]

    Data_Masses_Monopile_TransitionPiece = [
        ("Data_Masses_Monopile_TransitionPiece{" +
         f" id = \"{row['Name']}\", " +
         f" z =  {((row['Top [m]'] + row['Bottom [m]']) / 2):01.3f}, " +
         f" mass =  {row['Mass [kg]']:01.1f}" +
         "}") for _, row in ADDITIONAL_MASSES.iterrows()
    ]

    Data_Appurtenances = []
    for _, row in APPURTANCES.iterrows():
        parts = [
            f" id = \"{row['Name']}\"",
            f" z_bot = {row['Bottom [m]']:01.3f}",
            f" z_top = {row['Top [m]']:01.3f}",
            f" diameter = {row['Diameter [m]']:01.3f}",
            f" orientation = {row['Orientation [°]']:01.3f}",
            f" mass = {row['Mass [kg]']:01.1f}",
            f" surface_roughness = {row['Surface roughness [m]']:01.3f}"
        ]
        if pd.notna(row['Gap between surfaces [m]']):
            parts.insert(6, f" gap_between_surfaces = {row['Gap between surfaces [m]']}")
        if pd.notna(row['Distance Axis to Axis [m]']):
            parts.insert(6, f" distance_axis_to_axis = {row['Distance Axis to Axis [m]']}")
        Data_Appurtenances.append(" Data_Appurtenances{" + ", ".join(parts) + "}")

    Data_MarineGrowth = [
        ("Data_MarineGrowth{" +
         f" z_bot = \"{row['Bottom [m]']}\", " +
         f" z_top =  {row['Top [m]']:01.3f}, " +
         f" thickness =  {(row['Marine Growth [mm]'] / 1000):01.3f}, " +
         f" density =  {row['Density  [kg/m^3]']:01.0f}, " +
         f" surface_roughness =  {row['Surface Roughness [m]']:01.3f}" +
         "}") for _, row in MARINE_GROWTH.iterrows() if row['Marine Growth [mm]'] > 0
    ]

    text = (
            "--Input for WLGen generated by Excel\n\n"
            "--data for monopile sections: vertical positions z_bot and z_top are relative to design water level. Dimensions in meters.\n"
            + "\n".join(Data_MonopileSections) + "\n\n"
                                                 "--data for transition-piece sections: vertical positions z_bot and z_top are relative to design water level. Dimensions in meters.\n"
            + "\n".join(Data_TransitionPieceSections) + "\n\n"
                                                        "--data for additional masses: vertical positions z are relative to design water level.\n"
            + "\n".join(Data_Masses_Monopile_TransitionPiece) + "\n\n"
                                                                "--data for appurtenances: vertical positions z_bot and z_top are relative to design water level. Dimensions in meters, with the exception of orientation in degree from North and mass in kg.\n"
            + "\n".join(Data_Appurtenances) + "\n\n"
                                              "--data for marine growth: thickness up to defined vertical positions, relative to design water level, both in meters. The thickness is zero above the last given vertical position. Density in kg/m3.\n"
            + "\n".join(Data_MarineGrowth)
    )

    return text, ""


# %% macros
def export_JBOOST(excel_caller, jboost_path):
    excel_filename = os.path.basename(excel_caller)

    jboost_path = os.path.abspath(jboost_path)
    GEOMETRY = ex.read_excel_table(excel_filename, "StructureOverview", "WHOLE_STRUCTURE", dropnan=True)
    GEOMETRY = GEOMETRY.drop(columns=["Section", "local Section"])
    MASSES = ex.read_excel_table(excel_filename, "StructureOverview", "ALL_ADDED_MASSES", dropnan=True)
    MARINE_GROWTH = ex.read_excel_table(excel_filename, "StructureOverview", "MARINE_GROWTH", dropnan=True)
    PARAMETERS = ex.read_excel_table(excel_filename, "ExportStructure", "JBOOST_PARAMETER", dropnan=True)
    STRUCTURE_META = ex.read_excel_table(excel_filename, "StructureOverview", "STRUCTURE_META")
    PROJECT = ex.read_excel_table(excel_filename, "ExportStructure", "JBOOST_PROJECT", dtype=str)
    RNA = ex.read_excel_table(excel_filename, "StructureOverview", "RNA", dropnan=True)

    if len(RNA) == 0:
        ex.show_message_box(excel_filename,
                            f"Please define RNA parameters. Aborting")
        return

    # check Geometry
    sucess_GEOMETRY = mc.sanity_check_structure(excel_filename, GEOMETRY)
    if not sucess_GEOMETRY:
        ex.show_message_box(excel_filename,
                            f"Geometry is messed up. Aborting.")
        return

    Model_name = PARAMETERS.loc[PARAMETERS["Parameter"] == "ModelName", "Value"].values[0]

    # proj file
    PROJECT = PROJECT.set_index("Project Settings")

    default = PROJECT.loc[:, "default"]
    proj_configs = PROJECT.iloc[:, 2:]

    # iterate trough configs
    for config_name, config_data in proj_configs.items():

        # Fill missing values in config_data with defaults
        config_data = config_data.replace("", np.nan)
        config_data = config_data.fillna(default)

        # filling auto values
        def resolve_auto_value(parameter, config_key, description):
            var = STRUCTURE_META.loc[STRUCTURE_META["Parameter"] == parameter, "Value"].values
            if len(var) > 0 and isinstance(var[0], (int, float)):
                config_data[config_key] = var[0]
                return True
            else:
                ex.show_message_box(
                    excel_filename,
                    f"Please set {description} in the StructureOverview, as you set {config_key} in {config_name} to 'auto'. Aborting."
                )
                return False

        if config_data["water_level"] == 'auto':
            if not resolve_auto_value("Water level", "water_level", "a water level"):
                return
        if config_data["seabed_level"] == 'auto':
            if not resolve_auto_value("Seabed level", "seabed_level", "a seabed level"):
                return
        if config_data["h_hub"] == 'auto':
            if not resolve_auto_value("Hubheight", "h_hub", "Hubheight"):
                return
        if config_data["h_refwindspeed"] == 'auto':
            config_data["h_refwindspeed"] = config_data["h_hub"]

        config_struct = {row: data for row, data in config_data.items()}
        config_struct.pop("runFEModul", None)
        config_struct.pop("runFrequencyModul", None)

        def str_to_bool(s):
            if s == "True":
                return True
            elif s == "False":
                return False
            else:
                raise ValueError(f"Invalid boolean string: {s}.")

        runFEModul = str_to_bool(config_data["runFEModul"])
        runFrequencyModul = str_to_bool(config_data["runFrequencyModul"])

        proj_text = create_JBOOST_proj(config_struct,
                                       MARINE_GROWTH,
                                       modelname=Model_name,
                                       runFEModul=runFEModul,
                                       runFrequencyModul=runFrequencyModul,
                                       runHindcastValidation=False,
                                       wavefile="wave.lua",
                                       windfile="wind.")

        struct_text = create_JBOOST_struct(GEOMETRY,
                                           RNA,
                                           (PARAMETERS.loc[PARAMETERS["Parameter"] == "deflection MP", "Value"].values[
                                                0],
                                            PARAMETERS.loc[PARAMETERS["Parameter"] == "deflection MP", "Unit"].values[
                                                0]),
                                           (PARAMETERS.loc[
                                                PARAMETERS["Parameter"] == "deflection TOWER", "Value"].values[0],
                                            PARAMETERS.loc[
                                                PARAMETERS["Parameter"] == "deflection TOWER", "Unit"].values[0]),
                                           MASSES=MASSES,
                                           MARINE_GROWTH=MARINE_GROWTH,
                                           defl_TP=(PARAMETERS.loc[
                                                        PARAMETERS["Parameter"] == "deflection TP", "Value"].values[0],
                                                    PARAMETERS.loc[
                                                        PARAMETERS["Parameter"] == "deflection TP", "Unit"].values[0]),
                                           ModelName=Model_name,
                                           EModul=PARAMETERS.loc[PARAMETERS["Parameter"] == "EModul", "Value"].values[
                                               0],
                                           fyk="355",
                                           poisson="0.3",
                                           dens=
                                           PARAMETERS.loc[PARAMETERS["Parameter"] == "Steel Density", "Value"].values[
                                               0],
                                           addMass=0,
                                           member_id=1,
                                           create_node_tolerance=PARAMETERS.loc[PARAMETERS[
                                                                                    "Parameter"] == "Dimensional tolerance for node generating [m]", "Value"].values[
                                               0],
                                           seabed_level=STRUCTURE_META.loc[
                                               STRUCTURE_META["Parameter"] == "Seabed level", "Value"].values[0],
                                           waterlevel=config_struct["water_level"]
                                           )

        path_config = os.path.join(jboost_path, config_name)

        if not os.path.exists(path_config):
            os.mkdir(path_config)

        path_proj = os.path.join(path_config, config_name + ".lua")
        path_struct = os.path.join(path_config, Model_name + ".lua")

        with open(path_proj, 'w') as file:
            file.write(proj_text)
        with open(path_struct, 'w') as file:
            file.write(struct_text)

    ex.show_message_box(excel_filename,
                        f"JBOOST Structure {PARAMETERS.loc[PARAMETERS['Parameter'] == 'ModelName', 'Value'].values[0]} saved successfully at {jboost_path}")

    return


def export_WLGen(excel_caller, WLGen_path):
    excel_filename = os.path.basename(excel_caller)

    APPURTANCES_MASSES = ex.read_excel_table(excel_filename, "ExportStructure", "APPURTANCES")
    STRUCTURE = ex.read_excel_table(excel_filename, "StructureOverview", "WHOLE_STRUCTURE")
    STRUCTURE = STRUCTURE.drop(columns=["Section", "local Section"])
    MARINE_GROWTH = ex.read_excel_table(excel_filename, "StructureOverview", "MARINE_GROWTH")
    STRUCTURE_META = ex.read_excel_table(excel_filename, "StructureOverview", "STRUCTURE_META")
    SKIRT = ex.read_excel_table(excel_filename, "StructureOverview", "SKIRT", dropnan=True)

    APPURTANCES = APPURTANCES_MASSES.loc[APPURTANCES_MASSES.iloc[:, 0] == "WL", :]
    ADDITIONAL_MASSES = APPURTANCES_MASSES.loc[APPURTANCES_MASSES.iloc[:, 0] == "AM", :]

    # check APP

    # cut of below waterline
    z_wl = STRUCTURE_META.loc[STRUCTURE_META["Parameter"] == "Seabed level", "Value"].values[0]
    STRUCTURE = mc.add_element(STRUCTURE, z_new=z_wl)
    STRUCTURE = STRUCTURE[STRUCTURE["Bottom [m]"] >= z_wl]

    # only take MP and TP
    MP = STRUCTURE.loc[STRUCTURE.loc[:, "Affiliation"] == "MP", :]
    TP = STRUCTURE.loc[STRUCTURE.loc[:, "Affiliation"] == "TP", :]

    if len(SKIRT) == 0:
        SKIRT = None
    text, msg = create_WLGen_file(APPURTANCES, ADDITIONAL_MASSES, MP, TP, MARINE_GROWTH, skirt=SKIRT)

    # Feedback to user
    if not text:
        ex.show_message_box(excel_filename, f"WLGen Structure could not be created: {msg}")
    else:
        model_name = STRUCTURE_META.loc[STRUCTURE_META["Parameter"] == "Model Name", "Value"].values[0]
        if model_name is None:
            model_name = "WLGen_input.lua"
            ex.show_message_box(excel_filename,
                                f"No model name defined in Structure Overview. File named {model_name}.")
        else:
            model_name = model_name + ".lua"

        WLGen_path = os.path.abspath(os.path.join(WLGen_path, model_name))

        with open(WLGen_path, 'w') as file:
            file.write(text)
        ex.show_message_box(excel_filename, f"WLGen Structure created successfully and saved at {WLGen_path}.")
    return


def fill_WLGenMasses(excel_caller):
    excel_filename = os.path.basename(excel_caller)
    MASSES = ex.read_excel_table(excel_filename, "StructureOverview", "ALL_ADDED_MASSES")
    MASSES = MASSES.loc[(MASSES["Affiliation"] == "TP") | (MASSES["Affiliation"] == "MP")]

    def categorize_row(row):
        # Check mandatory fields
        if pd.isna(row['Top [m]']) or pd.isna(row['Mass [kg]']):
            return 'INVALID'

        # Check if it qualifies as an APPURTENANCE
        has_bottom = not pd.isna(row['Bottom [m]'])
        has_diameter = not pd.isna(row['Diameter [m]'])
        has_orientation = not pd.isna(row['Orientation [°]'])
        has_roughness = not pd.isna(row['Surface roughness [m]'])

        has_axis_to_axis = not pd.isna(row['Distance Axis to Axis [m]'])
        has_gap = not pd.isna(row['Gap between surfaces [m]'])
        # xor_axis_gap = has_axis_to_axis != has_gap  # exclusive OR

        if all([has_bottom, has_diameter, has_orientation, has_roughness]) and (has_axis_to_axis or has_gap):
            return 'WL'
        else:
            return 'AM'

    cols = ["Use For"] + list(MASSES.columns)
    MASSES_WL = pd.DataFrame(columns=cols)

    for idx, row in MASSES.iterrows():
        kind = categorize_row(row)

        row["Use For"] = kind

        row_df = row.to_frame().T  # Convert Series to 1-row DataFrame
        row_aligned = row_df[MASSES_WL.columns.intersection(row_df.columns)]

        MASSES_WL = pd.concat([MASSES_WL, row_aligned], ignore_index=True)

    # sort df
    # Define custom order
    cat_order = CategoricalDtype(categories=["WL", "AM", "INVALID"], ordered=True)

    # Convert the column to categorical
    MASSES_WL['Use For'] = MASSES_WL['Use For'].astype(cat_order)

    # Sort the DataFrame
    MASSES_WL = MASSES_WL.sort_values('Use For')

    ex.write_df_to_table(excel_filename, "ExportStructure", "APPURTANCES", MASSES_WL)


def fill_Bladed_table(excel_caller):
    excel_filename = os.path.basename(excel_caller)
    Bladed_Settings = ex.read_excel_table(excel_filename, "ExportStructure", "Bladed_Settings", dropnan=True)
    Bladed_Material = ex.read_excel_table(excel_filename, "ExportStructure", "Bladed_Material", dropnan=True)

    GEOMETRY = ex.read_excel_table(excel_filename, "StructureOverview", "WHOLE_STRUCTURE", dropnan=True)
    MARINE_GROWTH = ex.read_excel_table(excel_filename, "StructureOverview", "MARINE_GROWTH", dropnan=True)
    MASSES = ex.read_excel_table(excel_filename, "StructureOverview", "ALL_ADDED_MASSES")
    STRUCTURE_META = ex.read_excel_table(excel_filename, "StructureOverview", "STRUCTURE_META")

    Bladed_Elements = pd.DataFrame(
        columns=["Affiliation [-]", "Member [-]", "Node [-]", "Diameter [m]", "Wall thickness [mm]", "cd [-]", "cm [-]",
                 "Marine growth [mm]", "Density [kg*m^-3]", "Material [-]", "Elevation [m]"])
    Bladed_Nodes = pd.DataFrame(columns=["Node [-]", "Elevation [m]", "Local x [m]", "Local y [m]", "Point mass [kg]"])

    create_node_tolerance = Bladed_Settings.loc[
        Bladed_Settings["Parameter"] == "Dimensional Tolerance for Node generating [m]", "Value"].values[0]
    seabed_level = STRUCTURE_META.loc[STRUCTURE_META["Parameter"] == "Seabed level", "Value"].values[0]
    material = Bladed_Material.loc[0, "Material"]
    density = Bladed_Material.loc[0, "Density"]

    # Filter geometry below seabed
    if seabed_level is not None:
        GEOMETRY = mc.add_element(GEOMETRY, seabed_level)
    GEOMETRY = GEOMETRY.loc[GEOMETRY["Bottom [m]"] >= seabed_level]

    NODES = mc.extract_nodes_from_elements(GEOMETRY)

    # Add masses
    NODES["pMass"] = 0.0
    NODES["pMassNames"] = None
    NODES["added"] = False
    NODES["comment"] = None

    if MASSES is not None:
        for idx in MASSES.index:
            z_bot = MASSES.loc[idx, "Bottom [m]"]
            z_Mass = (z_bot + MASSES.loc[idx, "Top [m]"]) / 2 if pd.notna(z_bot) else MASSES.loc[idx, "Top [m]"]

            differences = np.abs(NODES["Elevation [m]"].values - z_Mass)
            within_tol = np.where(differences <= create_node_tolerance)[0]

            # if node is (nearly) on Node
            if len(within_tol) > 0:
                closest_index = within_tol[np.argmin(differences[within_tol])]
                NODES.loc[closest_index, "pMass"] += MASSES.loc[idx, "Mass [kg]"]

                if NODES.loc[closest_index, "comment"] is None:
                    NODES.loc[closest_index, "comment"] = MASSES.loc[idx, "Name"] + " "
                else:
                    NODES.loc[closest_index, "comment"] += MASSES.loc[idx, "Name"] + " "

            # if node mass is over bottom
            elif z_Mass >= GEOMETRY["Bottom [m]"].values[-1]:

                # add Node
                NODES = add_node(NODES, z_Mass, defaults={"float": 0})
                GEOMETRY = mc.add_element(GEOMETRY, z_Mass)

                NODES.loc[NODES["Elevation [m]"] == z_Mass, "added"] = True

                NODES.loc[NODES["Elevation [m]"] == z_Mass, "pMass"] += MASSES.loc[idx, "Mass [kg]"]

                NODES.loc[NODES["Elevation [m]"] == z_Mass, "comment"] = MASSES.loc[idx, "Name"] + " "

            else:
                print(f"Warning! Mass '{MASSES.loc[idx, 'Name']}' not added, it is below the seabed level!")

    # Nodes
    Bladed_Nodes.loc[:, "Node [-]"] = np.linspace(1, len(NODES), len(NODES))
    Bladed_Nodes.loc[:, "Elevation [m]"] = NODES.loc[:, "Elevation [m]"]
    Bladed_Nodes.loc[:, "Local x [m]"] = 0.0
    Bladed_Nodes.loc[:, "Local y [m]"] = 0.0
    Bladed_Nodes.loc[:, "Point mass [kg]"] = NODES.loc[:, "pMass"]
    Bladed_Nodes.loc[:, "Added"] = NODES.loc[:, "added"]
    Bladed_Nodes.loc[:, "Comment"] = NODES.loc[:, "comment"]

    # Geometry
    GEOMETRY.loc[:, "Section"] = np.linspace(1, len(GEOMETRY), len(GEOMETRY))
    Bladed_Elements.loc[:, "Affiliation [-]"] = np.array(
        [[aff_elem, aff_elem] for aff_elem in GEOMETRY["Affiliation"].values]).flatten()
    Bladed_Elements.loc[:, "Member [-]"] = np.array(
        [[f"{int(sec_elem)} (End 1)", f"{int(sec_elem)} (End 2)"] for sec_elem in GEOMETRY["Section"].values]).flatten()

    Bladed_Elements.loc[:, "Elevation [m]"] = np.array(
        [[row["Top [m]"], row["Bottom [m]"]] for i, row in GEOMETRY.iterrows()]).flatten()

    for i, row in Bladed_Elements.iterrows():

        # node
        elevation = row["Elevation [m]"]
        # Find matching node based on elevation
        node = Bladed_Nodes.loc[Bladed_Nodes["Elevation [m]"] == elevation, "Node [-]"]
        if not node.empty:
            Bladed_Elements.at[i, "Node [-]"] = int(node.values[0])

        marineGrowth = MARINE_GROWTH.loc[
            (MARINE_GROWTH["Bottom [m]"] < elevation) & (MARINE_GROWTH["Top [m]"] >= elevation), "Marine Growth [mm]"]

        if not marineGrowth.empty:
            Bladed_Elements.at[i, "Marine growth [mm]"] = marineGrowth.values[0]
        else:
            Bladed_Elements.at[i, "Marine growth [mm]"] = 0

    Bladed_Elements.drop(columns=["Elevation [m]"], inplace=True)

    Bladed_Elements.loc[:, "Diameter [m]"] = np.array(
        [[row["D, top [m]"], row["D, bottom [m]"]] for i, row in GEOMETRY.iterrows()]).flatten()
    Bladed_Elements.loc[:, "Wall thickness [mm]"] = np.array(
        [[row["t [mm]"], row["t [mm]"]] for i, row in GEOMETRY.iterrows()]).flatten()

    Bladed_Elements.loc[:, "cd [-]"] = 0.9
    Bladed_Elements.loc[:, "cm [-]"] = 2.0
    Bladed_Elements.loc[:, "Density [kg*m^-3]"] = density
    Bladed_Elements.loc[:, "Material [-]"] = material

    ex.write_df_to_table(excel_filename, "ExportStructure", "Bladed_Elements", Bladed_Elements)
    ex.write_df_to_table(excel_filename, "ExportStructure", "Bladed_Nodes", Bladed_Nodes)

    return


def fill_Sesam_table(excel_caller):
    excel_filename = os.path.basename(excel_caller)
    MASSES = ex.read_excel_table(excel_filename, "StructureOverview", "ALL_ADDED_MASSES")
    STRUCTURE_META = ex.read_excel_table(excel_filename, "StructureOverview", "STRUCTURE_META")
    GEOMETRY = ex.read_excel_table(excel_filename, "StructureOverview", "WHOLE_STRUCTURE", dropnan=True)

    GEOMETRY_formatted = build_GEOMETRY_formatted(GEOMETRY, STRUCTURE_META)
    ex.write_df_to_table(excel_filename, "ExportStructure", "tbl_ExportStructure_Structure",
                         GEOMETRY_formatted)

    MASSES_formatted = build_MASSES_formatted(MASSES, STRUCTURE_META)
    ex.write_df_to_table(excel_filename, "ExportStructure", "tbl_ExportStructure_Mass", MASSES_formatted)

    return


def build_GEOMETRY_formatted(GEOMETRY: pd.DataFrame,
                             STRUCTURE_META: pd.DataFrame,
                             density_kg_per_m3: float = 7850.0) -> pd.DataFrame:
    """
    Create the GEOMETRY_formatted dataframe used for export.
    Requires columns in GEOMETRY: 'Affiliation','Top [m]','Bottom [m]','D, top [m]','D, bottom [m]','t [mm]'.
    Requires STRUCTURE_META with row Parameter == 'Height Reference'.
    """
    # 1) height reference (no silent fallback)
    mask = STRUCTURE_META["Parameter"].str.strip().str.lower() == "height reference"
    if not mask.any():
        raise ValueError("Height Reference not found in STRUCTURE_META dataframe!")
    height_ref = STRUCTURE_META.loc[mask, "Value"].astype(str).iloc[0].strip()

    # 2) global running section number
    GEOMETRY = GEOMETRY.copy()
    GEOMETRY["section"] = range(1, len(GEOMETRY) + 1)

    # 3) label "TP / MP \nTower / Grout"
    prefix_map = {"TOWER": "TW", "TP": "TP", "TRANSITION PIECE": "TP", "MP": "MP", "MONOPILE": "MP"}
    prefix = GEOMETRY["Affiliation"].astype(str).str.strip().str.upper().map(prefix_map).fillna("XX")
    local_idx = (prefix.groupby(prefix).cumcount() + 1).astype(int).astype(str).str.zfill(2)
    label_header = "TP / MP \nTower / Grout"
    GEOMETRY[label_header] = prefix + local_idx

    # 4) rename to target headers
    rename_map = {
        "Top [m]": f"top [m {height_ref}]",
        "Bottom [m]": f"bottom [m {height_ref}]",
        "D, top [m]": "Diameter \ntop [m]",
        "D, bottom [m]": "Diameter bottom [m]",
    }
    G = GEOMETRY.rename(columns=rename_map)

    # 5) select/export order
    col_top = f"top [m {height_ref}]"
    col_bottom = f"bottom [m {height_ref}]"
    col_dt = "Diameter \ntop [m]"
    col_db = "Diameter bottom [m]"
    col_tmm = "t [mm]"
    cols_base = [label_header, "section", col_top, col_bottom, col_dt, col_db, col_tmm]
    GEOMETRY_formatted = G[cols_base].copy()

    # 6) numeric types + rounding for base cols
    round_spec = {col_top: 3, col_bottom: 3, col_dt: 3, col_db: 3, col_tmm: 0}
    for c, nd in round_spec.items():
        GEOMETRY_formatted[c] = pd.to_numeric(GEOMETRY_formatted[c], errors="coerce").round(nd)
    GEOMETRY_formatted["section"] = GEOMETRY_formatted["section"].astype(int)

    # 7) derived columns
    # l [m]
    GEOMETRY_formatted["l [m]"] = (pd.to_numeric(GEOMETRY_formatted[col_top], errors="coerce")
                                   - pd.to_numeric(GEOMETRY_formatted[col_bottom], errors="coerce"))
    # slope [°]
    num = pd.to_numeric(GEOMETRY_formatted[col_db], errors="coerce") - pd.to_numeric(GEOMETRY_formatted[col_dt],
                                                                                     errors="coerce")
    den = 2 * pd.to_numeric(GEOMETRY_formatted["l [m]"], errors="coerce")
    slope_deg = np.degrees(np.arctan(np.divide(num, den, out=np.zeros_like(num, dtype=float), where=den != 0)))
    slope_deg = np.where(den == 0, 0.0, slope_deg)
    GEOMETRY_formatted["slope [°]"] = slope_deg

    # weight [t]
    t_m = pd.to_numeric(GEOMETRY_formatted[col_tmm], errors="coerce") / 1000.0
    Dt = pd.to_numeric(GEOMETRY_formatted[col_dt], errors="coerce")
    Db = pd.to_numeric(GEOMETRY_formatted[col_db], errors="coerce")
    l = pd.to_numeric(GEOMETRY_formatted["l [m]"], errors="coerce")
    rho = float(density_kg_per_m3)

    vol_cyl_shell = (np.pi / 4.0) * (Dt ** 2 - (Dt - 2 * t_m) ** 2) * l
    outer = (Db ** 2 + Db * Dt + Dt ** 2)
    inner = ((Db - 2 * t_m) ** 2 + (Db - 2 * t_m) * (Dt - 2 * t_m) + (Dt - 2 * t_m) ** 2)
    vol_cone_shell = (l * np.pi / 12.0) * (outer - inner)

    w_cyl_t = vol_cyl_shell * rho / 1000.0
    w_cone_t = vol_cone_shell * rho / 1000.0
    GEOMETRY_formatted["weight [t]"] = np.where(np.isclose(GEOMETRY_formatted["slope [°]"], 0.0, atol=1e-12),
                                                w_cyl_t, w_cone_t)

    # SCF [-]
    t_curr = pd.to_numeric(GEOMETRY_formatted[col_tmm], errors="coerce")
    t_next = t_curr.shift(-1)
    t_min = np.minimum(t_curr, t_next)
    t_max = np.maximum(t_curr, t_next)
    with np.errstate(divide="ignore", invalid="ignore"):
        term = (np.abs(t_next - t_curr) / 2.0) / t_min
        blend = (t_min ** 1.5) / (t_min ** 1.5 + t_max ** 1.5)
        scf = 1.0 + 6.0 * term * blend
    scf = np.where(t_next.isna(), 1.0, scf)
    GEOMETRY_formatted["SCF [-]"] = scf

    # 8) final rounding
    GEOMETRY_formatted[col_top] = GEOMETRY_formatted[col_top].round(3)
    GEOMETRY_formatted[col_bottom] = GEOMETRY_formatted[col_bottom].round(3)
    GEOMETRY_formatted[col_dt] = GEOMETRY_formatted[col_dt].round(3)
    GEOMETRY_formatted[col_db] = GEOMETRY_formatted[col_db].round(3)
    GEOMETRY_formatted[col_tmm] = GEOMETRY_formatted[col_tmm].round(0)
    GEOMETRY_formatted["l [m]"] = GEOMETRY_formatted["l [m]"].round(3)
    GEOMETRY_formatted["slope [°]"] = GEOMETRY_formatted["slope [°]"].round(2)
    GEOMETRY_formatted["weight [t]"] = GEOMETRY_formatted["weight [t]"].round(2)
    GEOMETRY_formatted["SCF [-]"] = GEOMETRY_formatted["SCF [-]"].round(2)

    # 9) final column order
    GEOMETRY_formatted = GEOMETRY_formatted[[
        label_header, "section", col_top, col_bottom, col_dt, col_db, col_tmm,
        "l [m]", "slope [°]", "weight [t]", "SCF [-]"
    ]]

    return GEOMETRY_formatted


def build_MASSES_formatted(MASSES: pd.DataFrame, STRUCTURE_META: pd.DataFrame) -> pd.DataFrame:
    """Create MASSES_formatted with columns:
       Name | Height for Point mass [m <ref>] | Point mass [kg]
    """
    # --- height reference (no silent fallback)
    mask = STRUCTURE_META["Parameter"].str.strip().str.lower() == "height reference"
    if not mask.any():
        raise ValueError("Height Reference not found in STRUCTURE_META dataframe!")
    height_ref = STRUCTURE_META.loc[mask, "Value"].astype(str).iloc[0].strip()

    # --- normalize numeric inputs
    top = pd.to_numeric(MASSES.get("Top [m]"), errors="coerce")
    bottom = pd.to_numeric(MASSES.get("Bottom [m]"), errors="coerce")
    masskg = pd.to_numeric(MASSES.get("Mass [kg]"), errors="coerce")

    # height: if both present -> midpoint, else whichever exists
    height = np.where(~top.isna() & ~bottom.isna(), (top + bottom) / 2.0,
                      np.where(top.isna(), bottom, top))

    # keep only valid rows
    keep = (~pd.isna(height)) & (~pd.isna(masskg)) & (masskg != 0)
    df = MASSES.loc[keep].copy()
    height = pd.Series(height, index=MASSES.index).loc[keep]
    masskg = masskg.loc[keep]

    # --- name prefix by affiliation, counter per prefix
    prefix_map = {"TOWER": "TW", "TP": "TP", "TRANSITION PIECE": "TP", "MP": "MP", "MONOPILE": "MP"}
    prefix = df["Affiliation"].astype(str).str.strip().str.upper().map(prefix_map).fillna("XX")
    idx = (prefix.groupby(prefix).cumcount() + 1).astype(int).astype(str).str.zfill(2)
    name = prefix + "Mass" + idx

    # --- build formatted df
    col_height = f"Height for Point mass [m {height_ref}]"
    df_out = pd.DataFrame({
        "Name": name,
        col_height: height.round(3),
        "Point mass [kg]": masskg.round(0).astype("Int64")
    })

    # optional: sort by height descending (like your sheet)
    df_out = df_out.sort_values(by=col_height, ascending=False, kind="stable").reset_index(drop=True)

    return df_out



import os, re
import pandas as pd  # only used for the mass fallback

def export_Sesam(excel_caller, Sesam_path):
    excel_path = excel_caller
    excel_filename = os.path.basename(excel_caller)

    # --- model_name with safe default ---
    try:
        df_META = ex.read_excel_table(excel_filename, "StructureOverview", "STRUCTURE_META")
        mask = df_META["Parameter"].astype(str).str.strip().str.casefold().eq("model name")
        raw = None if df_META.loc[mask, "Value"].empty else df_META.loc[mask, "Value"].iloc[0]
        model_name = ("" if raw is None else str(raw)).strip()
    except Exception:
        model_name = ""

    if not model_name or model_name.lower() in {"nan", "none"}:
        model_name = "Sesam_input.js"
    else:
        if not model_name.lower().endswith(".js"):
            model_name += ".js"
        model_name = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", model_name).strip() or "Sesam_input.js"

    # --- preface from tbl_Export_Text (first column), fallback to BJ5:BJ11 ---
    preface_lines = []
    try:
        df_pref = ex.read_excel_table(excel_filename, "ExportStructure", "tbl_Export_Text", dropnan=False)
    except Exception:
        df_pref = None

    if df_pref is not None and df_pref.shape[0] > 0:
        col0 = df_pref.columns[0]
        preface_lines = [s for s in (("" if v is None else str(v).strip()) for v in df_pref[col0]) if s]
    else:
        vals = ex.read_excel_range(excel_path, "ExportStructure", "BJ5:BJ11")  # can be scalar/1D/2D
        # normalize to flat list of strings
        if vals is None:
            preface_lines = []
        elif isinstance(vals, (list, tuple)) and vals and isinstance(vals[0], (list, tuple)):
            preface_lines = [str(v).strip() for row in vals for v in row if v not in (None, "")]
        elif isinstance(vals, (list, tuple)):  # 1D
            preface_lines = [str(v).strip() for v in vals if v not in (None, "")]
        else:  # scalar
            s = str(vals).strip()
            preface_lines = [s] if s else []

    # --- main table ---
    try:
        df_Sesam = ex.read_excel_table(
            excel_filename, sheet_name="ExportStructure", table_name="tbl_Export_Sesam", dropnan=True
        )
    except Exception as e:
        ex.show_message_box(excel_filename, f"Failed to read 'tbl_Export_Sesam': {e}")
        return False
    if df_Sesam is None or len(df_Sesam) == 0:
        ex.show_message_box(excel_filename, "Table 'tbl_Export_Sesam' is empty.")
        return False

    # --- mass table with robust fallback ---
    df_Sesam_Mass = None
    try:
        df_Sesam_Mass = ex.read_excel_table(
            excel_filename, sheet_name="ExportStructure", table_name="tbl_Export_Sesam_Mass", dropnan=True
        )
    except Exception as e:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True, read_only=True)
            ws = wb["ExportStructure"]

            target = "point mass data"
            header_cell = None
            for r in ws.iter_rows(min_row=1, max_row=300, min_col=1, max_col=1000):
                for c in r:
                    if ("" if c.value is None else str(c.value)).strip().casefold() == target:
                        header_cell = c; break
                if header_cell: break
            if not header_cell:
                wb.close()
                ex.show_message_box(excel_filename,
                    "Failed to read 'tbl_Export_Sesam_Mass' and fallback could not find header 'Point Mass data'.")
                return False

            vals, empty_streak, row = [], 0, header_cell.row + 1
            while row <= ws.max_row and empty_streak < 50:
                v = ws.cell(row=row, column=header_cell.column).value
                if v is None or str(v).strip() == "":
                    empty_streak += 1
                else:
                    vals.append(str(v)); empty_streak = 0
                row += 1
            wb.close()

            df_Sesam_Mass = pd.DataFrame({"Point Mass data": vals})
        except Exception as e2:
            ex.show_message_box(excel_filename,
                f"Failed to read 'tbl_Export_Sesam_Mass' and fallback failed as well:\n{e}\n{e2}")
            return False

    # --- clean rows ---
    df_Sesam = df_Sesam.dropna(how="all")
    df_Sesam = df_Sesam[~(df_Sesam.astype(str).apply(lambda r: "".join(r), axis=1).str.strip() == "")].reset_index(drop=True)
    if df_Sesam_Mass is not None and len(df_Sesam_Mass) > 0:
        df_Sesam_Mass = df_Sesam_Mass.dropna(how="all")
        df_Sesam_Mass = df_Sesam_Mass[~(df_Sesam_Mass.astype(str).apply(lambda r: "".join(r), axis=1).str.strip() == "")].reset_index(drop=True)

    # --- output path ---
    try:
        out_path = os.path.abspath(os.path.join(Sesam_path, model_name))
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
    except Exception as e:
        ex.show_message_box(excel_filename, f"Output path error: {e}")
        return False

    # --- build & write file (ensure only strings go into 'lines') ---
    try:
        lines = []

        if preface_lines:
            lines.extend(preface_lines)   # list of strings
            lines.append("")

        lines.append("// ===== tbl_Export_Sesam =====")
        for col in df_Sesam.columns:
            lines.append(f"// {col}")
            for v in df_Sesam[col]:
                s = "" if v is None else str(v).strip()
                if s: lines.append(s)
            lines.append("")

        if df_Sesam_Mass is not None and len(df_Sesam_Mass) > 0:
            lines.append("// ===== tbl_Export_Sesam_Mass =====")
            for col in df_Sesam_Mass.columns:
                lines.append(f"// {col}")
                for v in df_Sesam_Mass[col]:
                    s = "" if v is None else str(v).strip()
                    if s: lines.append(s)
                lines.append("")

        # safety: stringify everything
        lines = ["" if x is None else str(x) for x in lines]

        with open(out_path, "w", encoding="utf-8", newline="\n") as f:
            f.write("\n".join(lines))

        ex.show_message_box(
            excel_filename,
            f"Sesam file created:\n{out_path}\n\n"
            f"Rows (tbl_Export_Sesam): {len(df_Sesam)} \n"
            f"Rows (tbl_Export_Sesam_Mass): {0 if df_Sesam_Mass is None else len(df_Sesam_Mass)}\n"
            f"Preface lines (tbl_Export_Text): {len(preface_lines)}"
        )
        return True
    except Exception as e:
        ex.show_message_box(excel_filename, f"Sesam file could not be created: {e}")
        return False




if __name__ == "__main__":
    export_Sesam(
        r"C:\temp\tutorial\v1.0\GeometryConverter.xlsm", r"C:/temp/tutorial/output/"
    )
